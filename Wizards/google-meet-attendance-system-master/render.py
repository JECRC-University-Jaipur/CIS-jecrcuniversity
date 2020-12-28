import eel
from pyjson import whconfig, write_data
import json
from testing import attendance
from datetime import datetime
from datetime import time
from datetime import date
import openpyxl as xl
from threading import Timer
import os

eel.init("gui")
timer = []


@eel.expose
def get_directory():
    print("here")
    from tkinter import filedialog
    from tkinter import Tk

    root = Tk()
    root.withdraw()
    root.call("wm", "attributes", ".", "-topmost", True)
    folder_selected = filedialog.askdirectory()
    if whconfig["output"] == "":
        if not os.path.exists(os.path.join(folder_selected, "screenshots")):
            os.mkdir(os.path.join(folder_selected, "screenshots"))
        if not os.path.exists(os.path.join(folder_selected, "excel")):
            os.mkdir(os.path.join(folder_selected, "excel"))
        whconfig["image_location"] = os.path.join(folder_selected, "screenshots")
        whconfig["excel_location"] = os.path.join(folder_selected, "excel")
    else:
        if not os.path.exists(os.path.join(whconfig["output"], "screenshots")):
            os.mkdir(os.path.join(whconfig["output"], "screenshots"))
        if not os.path.exists(os.path.join(whconfig["output"], "excel")):
            os.mkdir(os.path.join(whconfig["output"], "excel"))
        whconfig["image_location"] = os.path.join(whconfig["output"], "screenshots")
        whconfig["excel_location"] = os.path.join(whconfig["output"], "excel")
    print(folder_selected)
    return folder_selected


@eel.expose
def returnJsonData(*args):
    if len(args) > 0:
        try:
            return json.dumps(whconfig[args[0]])
        except KeyError as e:
            return 0
    return json.dumps(whconfig)


@eel.expose
def changeJsonData():
    data = eel.sendFormData()()
    print(json.loads(data))
    for key, value in json.loads(data).items():
        whconfig[key] = value
    print(data)
    print(whconfig)
    write_data(whconfig)


@eel.expose
def start_timers():
    while len(whconfig["times"]) > 0:
        checkTime = datetime.now().strftime("%H:%M")
        if checkTime in whconfig["times"]:
            whconfig["meeting_data"] = []
            write_data(whconfig)
            whconfig["times"].pop(whconfig["times"].index(checkTime))
            print(whconfig["times"])
            write_data(whconfig)
            attendance()


@eel.expose
def save_excel(data):
    if os.path.exists(os.path.join(whconfig["excel_location"], whconfig["excel_name"])):
        wb = xl.load_workbook(
            os.path.join(whconfig["excel_location"], whconfig["excel_name"])
        )
        ws = wb.active
    else:
        wb = xl.Workbook()
        ws = wb.active
    col = ws.max_column + 1
    print(data)
    for i in range(1, len(data) + 1):
        ws.cell(row=i, column=col).value = data[i - 1]
    wb.save(os.path.join(whconfig["excel_location"], whconfig["excel_name"]))
    wb.close()
    return


if not os.path.exists(os.path.join(whconfig["driver_path"])):
    os.mkdir(os.path.join(whconfig["driver_path"]))
eel.start(
    "index.html",
    size=(650, 612),
    render_template="jinja",
    cmdline_args=["--disable-dev-tools"],
)
